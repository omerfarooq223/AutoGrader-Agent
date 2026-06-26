"""
Excel writer utility — generates the final grading report with:
  - Main grading sheet (Name, ID, Marks, Category Scores, Deductions, Plagiarism)
  - Summary statistics sheet (avg, median, min, max, pass/fail, grade distribution)
  - Class insights section (top 3 common mistakes via LLM analysis)
"""

import logging
import re
import statistics

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import PASS_THRESHOLD, TOTAL_MARKS

logger = logging.getLogger(__name__)


# ── Styling constants ───────────────────────────────────────────
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
_SUBHEADER_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
_PASS_FILL   = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
_PASS_FONT   = Font(color="166534", bold=True)
_FAIL_FILL   = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
_FAIL_FONT   = Font(color="9F1239", bold=True)

_FLAG_FONT   = Font(color="B91C1C", bold=True)
_FLAG_FILL   = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),  bottom=Side(style="thin", color="CBD5E1"),
)
_ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
_BODY_FONT = Font(color="0F172A", size=10)
_MUTED_FONT = Font(color="475569", size=10)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TOP = Alignment(vertical="top", wrap_text=True)


def _shorten_flag(flag: str) -> str:
    """
    Convert a verbose plagiarism flag string into a readable summary that
    preserves the names of matched students so teachers know who copied from whom.

    Input:  'Similar to ali.pdf (95.7%, cos=96% ngram=96%) | Similar to sara.docx (...)'
    Output: '⚠️ Similar to: ali.pdf (96%), sara.docx (96%)'
    """
    if not flag:
        return ""
    matches = [m.strip() for m in flag.split("|") if m.strip()]
    if not matches:
        return ""
    parts = []
    for m in matches:
        # Extract filename and percentage from "Similar to X.pdf (95.7%, ...)"
        name_match = re.search(r"Similar to (.+?) \(", m)
        pct_match  = re.search(r"\((\d+(?:\.\d+)?)%", m)
        if name_match and pct_match:
            parts.append(f"{name_match.group(1)} ({float(pct_match.group(1)):.0f}%)")
        else:
            parts.append(m)
    return "⚠️ Similar to: " + ", ".join(parts)


def shorten_plagiarism_flag(flag: str) -> str:
    """
    Public wrapper used by the UI to shorten plagiarism flag strings.
    """
    if flag is None:
        return ""
    if not isinstance(flag, str):
        try:
            flag = str(flag)
        except Exception:
            return ""
    # Common pandas missing value representation
    if flag.strip().lower() == "nan":
        return ""
    return _shorten_flag(flag)


def _clean_category_name(name: str) -> str:
    """Strip leading/trailing brackets from category names."""
    return name.strip("[]")


def _generate_class_insights(results: list[dict]) -> list[str]:
    """
    Deterministically infer top 3 common mistakes from deduction text.
    This avoids LLM dependency for class insights generation.
    """
    parsed: list[tuple[str, str]] = []
    for r in results:
        d = r.get("deductions", "") or ""
        if not d or d == "No deductions.":
            continue
        for crit, reason, _ded in re.findall(
            r"([^:]+):\s*([^(,;|]+?)\s*\(-\s*(\d+(?:\.\d+)?)\)",
            d,
        ):
            parsed.append((crit.strip(), reason.strip().lower()))

    if not parsed:
        return []

    by_criterion: dict[str, int] = {}
    by_reason: dict[str, int] = {}
    for crit, reason in parsed:
        by_criterion[crit] = by_criterion.get(crit, 0) + 1
        if reason and reason not in {"marks deducted", "criterion not evaluated by grader"}:
            by_reason[reason] = by_reason.get(reason, 0) + 1

    top_criteria = sorted(by_criterion.items(), key=lambda x: x[1], reverse=True)
    top_reasons = sorted(by_reason.items(), key=lambda x: x[1], reverse=True)

    insights: list[str] = []
    for crit, count in top_criteria[:2]:
        insights.append(f"{crit} was the most commonly deducted criterion ({count} submission(s)).")
    if top_reasons:
        reason, count = top_reasons[0]
        insights.append(f"Most repeated deduction reason: {reason} ({count} submission(s)).")

    return insights[:3]


def generate_class_insights(results: list[dict]) -> list[str]:
    """
    Public wrapper so the UI can render the same "Class Insights" section
    without duplicating logic.
    """
    return _generate_class_insights(results)


def _auto_width(ws) -> None:
    """Auto-fit column widths based on cell contents."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        is_number_col = True
        for cell in col:
            cell.alignment = _TOP
            if cell.value:
                val = cell.value
                if not isinstance(val, (int, float)):
                    is_number_col = False
                max_len = max(max_len, len(str(val)))
        max_width = 15 if is_number_col else 60
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = max(
            ws.row_dimensions[row[0].row].height or 0, 30
        )


def _add_table(ws, table_name: str, last_row: int, last_col: int, style: str) -> None:
    """Apply an Excel table style over the populated area."""
    if last_row < 2 or last_col < 1:
        return
    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _add_score_bars(ws, start_row: int, end_row: int, columns: list[int]) -> None:
    """Add subtle data bars to numeric score columns."""
    if end_row < start_row:
        return
    for col in columns:
        letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{letter}{start_row}:{letter}{end_row}",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="max",
                color="60A5FA",
                showValue=True,
            ),
        )


def _collect_all_categories(results: list[dict]) -> list[str]:
    """Gather a sorted union of all category names across results."""
    cats: set[str] = set()
    for r in results:
        cats.update(r.get("category_scores", {}).keys())
    return sorted(cats)


def _write_grading_sheet(wb: openpyxl.Workbook, results: list[dict], sheet_title: str = "") -> None:
    """Write the main Grading Report sheet."""
    ws = wb.active
    ws.title = (sheet_title[:31] if sheet_title else "Grading Report")  # Excel 31 char limit

    categories = _collect_all_categories(results)

    # Build headers — no brackets on category names, no Feedback column
    # Derive total from max observed score — accurate regardless of config TOTAL_MARKS
    numeric_marks = [r.get("marks") for r in results if isinstance(r.get("marks"), (int, float))]
    rubric_total = int(max(numeric_marks)) if numeric_marks else (TOTAL_MARKS if TOTAL_MARKS else "?")
    pass_mark = round(rubric_total * (PASS_THRESHOLD / 100.0)) if isinstance(rubric_total, (int, float)) else 0
    headers = ["Name", "ID", f"Marks (/ {rubric_total})"]
    headers += [_clean_category_name(c) for c in categories]
    headers += ["Deductions / Reason", "Plagiarism Flag"]   # Feedback removed

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _THIN_BORDER
    ws.row_dimensions[1].height = 36

    # Data rows
    for row_idx, entry in enumerate(results, start=2):
        col = 1

        name_cell = ws.cell(row=row_idx, column=col, value=entry.get("name", ""))
        name_cell.font = _BODY_FONT
        col += 1
        id_cell = ws.cell(row=row_idx, column=col, value=entry.get("id", ""))
        id_cell.font = _MUTED_FONT
        id_cell.alignment = _CENTER
        col += 1

        # Marks with pass/fail colouring
        marks = entry.get("marks", "")
        marks_cell = ws.cell(row=row_idx, column=col, value=marks)
        marks_cell.alignment = _CENTER
        col += 1
        if isinstance(marks, (int, float)):
            if marks >= pass_mark:
                marks_cell.fill = _PASS_FILL
                marks_cell.font = _PASS_FONT
            else:
                marks_cell.fill = _FAIL_FILL
                marks_cell.font = _FAIL_FONT

        # Category scores — look up by original key (may have brackets)
        cat_scores = entry.get("category_scores", {})
        for cat in categories:
            # Try exact key first, then stripped version for bracket mismatches
            value = cat_scores.get(cat)
            if value is None:
                stripped = cat.strip("[]")
                value = cat_scores.get(stripped, "")
            score_cell = ws.cell(row=row_idx, column=col, value=value)
            score_cell.alignment = _CENTER
            score_cell.font = _BODY_FONT
            col += 1

        # Deductions — now built deterministically by Python in grader_agent
        deductions = entry.get("deductions", "") or "No deductions."
        deduction_cell = ws.cell(row=row_idx, column=col, value=deductions)
        deduction_cell.font = _BODY_FONT
        deduction_cell.alignment = _TOP
        col += 1

        # Plagiarism flag — shortened
        raw_flag  = entry.get("plagiarism_flag", "")
        short_flag = _shorten_flag(raw_flag)
        flag_cell  = ws.cell(row=row_idx, column=col, value=short_flag)
        flag_cell.alignment = _TOP
        if short_flag:
            flag_cell.font = _FLAG_FONT
            flag_cell.fill = _FLAG_FILL
        else:
            flag_cell.font = _MUTED_FONT

        # Borders and zebra striping
        is_even = (row_idx % 2 == 0)
        for c in range(1, col + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = _THIN_BORDER
            # Apply zebra striping ONLY if the cell doesn't already have a fill
            if is_even and cell.fill.fill_type is None:
                cell.fill = _ALT_ROW_FILL

    last_row = len(results) + 1
    last_col = len(headers)
    score_columns = [3] + list(range(4, 4 + len(categories)))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "2563EB"
    _add_table(ws, "GradingReportTable", last_row, last_col, "TableStyleMedium2")
    _add_score_bars(ws, 2, last_row, score_columns)
    _auto_width(ws)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 24)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 16)
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions[get_column_letter(last_col - 1)].width = 70
    ws.column_dimensions[get_column_letter(last_col)].width = 50
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        for idx in score_columns:
            row[idx - 1].alignment = _CENTER


def _write_stats_sheet(wb: openpyxl.Workbook, results: list[dict],
                       assignment_name: str = "", course_code: str = "",
                       semester: str = "") -> tuple:
    """Write a Summary Statistics sheet. Returns (worksheet, last_row)."""
    ws = wb.create_sheet("Summary Statistics")

    marks         = [r["marks"] for r in results if isinstance(r.get("marks"), (int, float))]
    total_students = len(results)
    error_entries  = [r for r in results if r.get("marks") == "Error"]

    stats_data = []
    if assignment_name:
        label = assignment_name
        if course_code: label = f"{course_code} — {label}"
        if semester:    label = f"{label} ({semester})"
        stats_data.append(("Assignment", label))
        stats_data.append(("", ""))
    stats_data += [
        ("Total Submissions",      total_students),
        ("Graded (numeric marks)", len(marks)),
        ("Grading errors",         len(error_entries)),
    ]

    # List filenames that failed
    if error_entries:
        stats_data.append(("", ""))
        stats_data.append(("Failed Submissions", "Filename"))
        for r in error_entries:
            stats_data.append(("", r.get("filename", r.get("name", "unknown"))))

    if marks:
        # Pass threshold based on PASS_THRESHOLD percentage
        rubric_total  = int(max(marks))  # Derive from data — accurate regardless of config
        pass_mark     = round(rubric_total * (PASS_THRESHOLD / 100.0))
        passed        = sum(1 for m in marks if m >= pass_mark)
        stats_data += [
            ("", ""),
            ("Total Marks (per assignment)", rubric_total),
            ("Pass Mark",                   f"≥ {pass_mark} ({PASS_THRESHOLD}%)"),
            ("Average",       round(statistics.mean(marks), 2)),
            ("Median",        round(statistics.median(marks), 2)),
            ("Std Deviation", round(statistics.stdev(marks), 2) if len(marks) > 1 else "N/A"),
            ("Minimum",       min(marks)),
            ("Maximum",       max(marks)),
            ("", ""),
            (f"Passed (≥ {pass_mark})",  passed),
            (f"Failed (< {pass_mark})",  len(marks) - passed),
            ("Pass Rate",     f"{passed / len(marks) * 100:.1f}%"),
        ]

        # rubric_total already calculated above — use it here too
        buckets = {
            f"A (≥90% of {rubric_total})":   0,
            f"B (80-89% of {rubric_total})":  0,
            f"C (70-79% of {rubric_total})":  0,
            f"D (60-69% of {rubric_total})":  0,
            f"F (<60% of {rubric_total})":    0,
        }
        bk = list(buckets.keys())
        for m in marks:
            pct = (m / rubric_total) * 100
            if   pct >= 90: buckets[bk[0]] += 1
            elif pct >= 80: buckets[bk[1]] += 1
            elif pct >= 70: buckets[bk[2]] += 1
            elif pct >= 60: buckets[bk[3]] += 1
            else:           buckets[bk[4]] += 1

        stats_data.append(("", ""))
        stats_data.append(("Grade Distribution", "Count"))
        for grade, count in buckets.items():
            stats_data.append((grade, count))

    import datetime
    stats_data.append(("", ""))
    stats_data.append(("Report Metadata", ""))
    stats_data.append(("Grading Engine", "AutoGrader Agent"))
    stats_data.append(("Generated At", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    # Write headers
    for col_idx, h in enumerate(["Metric", "Value"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER
    ws.row_dimensions[1].height = 34

    for row_idx, (metric, value) in enumerate(stats_data, start=2):
        c1 = ws.cell(row=row_idx, column=1, value=metric)
        c2 = ws.cell(row=row_idx, column=2, value=value)
        c1.border = _THIN_BORDER
        c2.border = _THIN_BORDER
        c1.font = _BODY_FONT
        c2.font = _BODY_FONT
        c1.alignment = _TOP
        c2.alignment = _TOP

        if metric in ("Failed Submissions", "Grade Distribution", "Report Metadata"):
            c1.font = _HEADER_FONT
            c1.fill = _SUBHEADER_FILL
            c2.font = _HEADER_FONT
            c2.fill = _SUBHEADER_FILL
            c1.alignment = _CENTER
            c2.alignment = _CENTER
        elif row_idx % 2 == 0:
            c1.fill = _ALT_ROW_FILL
            c2.fill = _ALT_ROW_FILL

    last_row = len(stats_data) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{last_row}"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0F766E"
    _add_table(ws, "SummaryStatisticsTable", last_row, 2, "TableStyleMedium4")
    _auto_width(ws)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 34)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 42)
    return ws, last_row


def _write_insights_section(ws, start_row: int, insights: list[str]) -> None:
    """Append a 'Class Insights' section to the stats sheet."""
    if not insights:
        return

    _INSIGHT_FILL = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )

    row = start_row + 2  # blank separator
    cell = ws.cell(row=row, column=1, value="Class Insights — Top 3 Common Mistakes")
    cell.font   = Font(bold=True, size=11, color="92400E")
    cell.fill   = _INSIGHT_FILL
    cell.border = _THIN_BORDER
    cell.alignment = _CENTER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=2).border = _THIN_BORDER
    ws.cell(row=row, column=2).fill   = _INSIGHT_FILL

    for i, insight in enumerate(insights, start=1):
        row += 1
        idx_cell = ws.cell(row=row, column=1, value=f"#{i}")
        insight_cell = ws.cell(row=row, column=2, value=insight)
        idx_cell.border = _THIN_BORDER
        insight_cell.border = _THIN_BORDER
        idx_cell.font = Font(bold=True, color="92400E")
        idx_cell.alignment = _CENTER
        insight_cell.font = _BODY_FONT
        insight_cell.alignment = _TOP


from typing import Union

def write_results(
    results: list[dict],
    output_path: str = "results.xlsx",
    return_insights: bool = False,
    assignment_name: str = "",
    course_code: str = "",
    semester: str = "",
) -> Union[str, tuple[str, list[str]]]:
    """
    Write grading results to an Excel file with two sheets:
      1. Grading Report   — per-student results (no Feedback column)
      2. Summary Statistics — class-level stats + class insights

    Returns the path to the written file.
    """
    wb = openpyxl.Workbook()
    _write_grading_sheet(wb, results, sheet_title=assignment_name or "Grading Report")
    ws_stats, last_row = _write_stats_sheet(wb, results,
                                            assignment_name=assignment_name,
                                            course_code=course_code,
                                            semester=semester)

    insights = generate_class_insights(results)
    if insights:
        _write_insights_section(ws_stats, last_row, insights)
    _auto_width(ws_stats)

    # Atomic write — save to temp file first, then rename
    # Prevents corrupted output if process is killed mid-write
    import os
    import tempfile
    from pathlib import Path
    out = Path(output_path)
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", dir=out.parent, delete=False
    ) as tmp:
        tmp_path = tmp.name
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, output_path)
    except Exception:
        Path(tmp_path).unlink(missing_ok=True)
        raise
    if return_insights:
        return output_path, insights
    return output_path
